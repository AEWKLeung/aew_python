# haz_waste_check.py


import pandas as pd
import numpy as np
from io import BytesIO
from flask import send_file

### Helper Functions ###

def format_to_floats(str_int_or_float):
    string = str(str_int_or_float)
    clean_string = string.strip('< > *')
    if string == "NEG":
        return string
    if string == "POS":
        return string
    if '@' in string:
        return string
    else:
        return float(clean_string)

# Drop duplicates keeps row with most recent prepdate
# (for when samples are rerun with new standards)
def drop_duplicates(lab_df):
    lab_df.sort_values(by=['PREPDATE'], ascending=True)
    no_dupes = lab_df.drop_duplicates(subset=['SAMPID','ANALYTE'], keep='first')
    if len(no_dupes) != len(lab_df):
        print(f'WARNING {len(lab_df) - len(no_dupes)} duplicate samples were found and dropped.')
        print(pd.concat([no_dupes,lab_df]).drop_duplicates(keep=False))
        return no_dupes
    else:
        return lab_df

def format_results_with_rls(row):
    if row["FINALVAL"] == "ND":
        # rounding to 6 decimal places prevents trailing 0s
        return f'<{round(row["PQL"],6)}'
    if row["FINALVAL"] == "neg" or row["FINALVAL"]=='NEG':
        return "NEG"
    if row["FINALVAL"] == "pos" or row["FINALVAL"]=='POS':
        return "POS"
    # for asbestos results
    ### note this line breaks sometimes when looking at an integer
    if '<' in str(row["FINALVAL"]):
        return row["FINALVAL"]
    if '@' in str(row['FINALVAL']):
        return row['FINALVAL']
    else:
        return float(row["FINALVAL"])
    

def calculate_hazardous_results(row):
    result = format_to_floats(row['Results'])
    STLC = np.nan if row['STLC']=='NA' else row['STLC']
    TCLP = np.nan if row['TCLP']=='NA' else row['TCLP'] 
    TTLC = np.nan if row['TTLC']=='NA' else row['TTLC'] 
    # print(result, STLC, TCLP, TTLC) # for debugging purposes
    # If analyte has POS/NEG result
    if result == "NEG" or result == "POS":
        return('No Add-On Analysis, Non-Hazardous Waste')

        # If chemical has STLC and TCLP
        if result < STLC * 10 and result < TCLP * 20 and result < TTLC:
            return("No Add-On Analysis, Non-Hazardous Waste")
        if result >= STLC * 10 and result < TCLP * 20:
            return("Run WET Add-On Analysis")
        if result >= TCLP * 20 and result < TTLC:
            return("Run TCLP & WET Add-On Analysis")
        if result >= TCLP * 20 and result >= TTLC:
            return("Run TCLP Add-On Analysis, automatically Non-RCRA Hazardous")
        
        # If chemical only has a STLC
        if pd.isnull(TCLP):
            if result < STLC * 10 and result < TTLC:
                return("No Add-On Analysis, Non-Hazardous Waste")
            if result >= STLC * 10:
                return("Run WET Add-On Analysis")
        
        # If chemical only has a TCLP
        if result < TCLP * 20 and pd.isnull(STLC) and pd.isnull(TTLC):
            return("No Add-On Analysis, Non-Hazardous Waste")
        if result >= TCLP * 20 and pd.isnull(STLC) and pd.isnull(TTLC):
            return("Run TCLP Add-On Analysis")
        else:
            return("OTHER RESULT")

def generate_hazardous_waste_reports(lab_data_df,regulatory_data_df):
    lab_data = lab_data_df
    regulatory_data = regulatory_data_df.drop(index=0).reset_index(drop=True)    # drop category data, breaks downstream processing

    # Ignores chemicals that are not results (i.e. surrogates)
    lab_data = lab_data.dropna(subset=['PQL']) 

    # Drop duplicates, keep row with most recent prepdate
    # (for when samples are rerun with new standards)
    lab_data = drop_duplicates(lab_data) 

    # If EDD has column titled FINALVALUE, make FINALVAL
    lab_data.rename(columns = {'FINALVALUE':'FINALVAL'}, inplace = True)
    
    # Add results column with RLs (Change NDs to RLs)
    lab_data["Results"] = lab_data.apply(format_results_with_rls, axis="columns", result_type='reduce')

    ### Merge Results with Haz Waste Criteria ###
    lab_data_with_limits = pd.merge(
        lab_data, regulatory_data, how="left", left_on="ANALYTE", right_on="Reference",
    )

    # make a table with only relevant lab data and Haz Waste Criteria (TTLC, STLCx10, STLC, TCLPx20, TCLP)
    # may need to use "reindex", to order properly
    lab_data_haz_check = lab_data_with_limits[
        [
            "PROJNAME",
            "LABSAMPID",
            "SAMPID",
            "ANALYTE",
            "Results",
            "TTLC",
            "STLCx10",
            "STLC",
            "TCLPx20",
            "TCLP",
        ]
    ]  

    lab_data_haz_check['Analysis'] = lab_data_haz_check.apply(calculate_hazardous_results, axis=1)

    # Filter out results that don't have a Haz Waste Criteria
    haz_check_results_filtered = lab_data_haz_check[
        lab_data_haz_check["Analysis"] != "OTHER RESULT"
    ]

    # rename columns for output
    haz_check_results_filtered.rename(
        columns={
            "PROJNAME": "Project Name",
            "LABSAMPID": "Lab Sample ID",
            "SAMPID": "Sample ID",
            "Results": "Result (mg/Kg)",
            "ANALYTE": "Analyte",
        },
        inplace=True,
    )

    haz_waste_analysis = haz_check_results_filtered
    ### Genrate report where  Add-Ons Needed
    haz_add_ons = haz_check_results_filtered[
        haz_check_results_filtered["Analysis"] != "No Add-On Analysis, Non-Hazardous Waste"
    ]
    

    return haz_add_ons, haz_waste_analysis



from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.formatting.rule import FormatObject, CellIsRule, FormulaRule
import openpyxl

def format_and_export_haz_waste_analysis(haz_waste_analysis):
    wb = Workbook()
    ws = wb.active
    # Open DF in openpyxl 
    for r in dataframe_to_rows(haz_waste_analysis, index=False, header=True):
        ws.append(r)


    # Widen columns for readability and pretyyness
    for column in range(1, 6):
        column_letter = get_column_letter(column)
        ws.column_dimensions[column_letter].width = 18

        # Set  font for readaiblity and prettyness
        for row in range(1, len(haz_waste_analysis)+2):

            cell_font = Font(
                name='Calibri',
                size=12,
                bold=False,
                italic=False,
                underline='none',
                color='FF000000'
            )
            ws[f'{column_letter}{row}'].font = cell_font
            # ws[f'{column}{row}'].fill = 

    # Set alignment for number values
    for column in range(5,11):
        column_letter = get_column_letter(column)
        for row in range(1, len(haz_waste_analysis)+2):
            cell = ws[f'{column_letter}{row}']
            alignment_obj = cell.alignment.copy(horizontal='right', vertical='center', wrap_text=False)
            cell.alignment = alignment_obj 

    # Wrap, center, resize header for readaiblity and prettyness
    for cell in ws['1:1']:
        alignment_obj = cell.alignment.copy(horizontal='center', vertical='center', wrap_text=True)
        cell.alignment = alignment_obj 
        header_font = Font(
            name='Calibri',
            size=13,
            bold=True,
            italic=False,
            underline='none',
            color='FF000000'
        )
        cell.font = header_font

    # Highlight Exceedances 
    exceedance_fill = PatternFill(
        start_color='EE1111',
        end_color='EE1111',
        fill_type='solid'
    )
    value_col = 'E'

    ws.conditional_formatting.add(
        f'{value_col}2:{value_col}{len(haz_waste_analysis)+1}',
        FormulaRule(formula=['AND(ISNUMBER(E2), E2>=G2)'],  fill=exceedance_fill)
    )

    # Save File
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def format_and_export_haz_waste_add_on(haz_waste_add_ons):
    if len(haz_waste_add_ons) == 0:
        return "There are No Potential Add-Ons"
    wb = Workbook()
    ws = wb.active
    # Open DF in openpyxl 
    for r in dataframe_to_rows(haz_waste_add_ons, index=False, header=True):
        ws.append(r)


    # Widen columns for readability and pretyyness
    for column in range(1, 6):
        column_letter = get_column_letter(column)
        ws.column_dimensions[column_letter].width = 18

        # Set  font for readaiblity and prettyness
        for row in range(1, len(haz_waste_add_ons)):

            cell_font = Font(
                name='Calibri',
                size=12,
                bold=False,
                italic=False,
                underline='none',
                color='FF000000'
            )
            ws[f'{column_letter}{row}'].font = cell_font
            # ws[f'{column}{row}'].fill = 

    # Set alignment for number values
    for column in range(5,11):
        column_letter = get_column_letter(column)
        for row in range(1, len(haz_waste_add_ons)):
            cell = ws[f'{column_letter}{row}']
            alignment_obj = cell.alignment.copy(horizontal='right', vertical='center', wrap_text=False)
            cell.alignment = alignment_obj 

    # Wrap, center, resize header for readaiblity and prettyness
    for cell in ws['1:1']:
        alignment_obj = cell.alignment.copy(horizontal='center', vertical='center', wrap_text=True)
        cell.alignment = alignment_obj 
        header_font = Font(
            name='Calibri',
            size=13,
            bold=True,
            italic=False,
            underline='none',
            color='FF000000'
        )
        cell.font = header_font

    # Highlight Exceedances 
    exceedance_fill = PatternFill(
        start_color='EE1111',
        end_color='EE1111',
        fill_type='solid'
    )
    value_col = 'E'

    ws.conditional_formatting.add(
        f'{value_col}2:{value_col}{len(haz_waste_add_ons)+1}',
        FormulaRule(formula=['AND(ISNUMBER(E2), E2>=G2)'],  fill=exceedance_fill)
    )
    # Save File
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output




