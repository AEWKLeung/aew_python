
from gc import collect
from lib2to3.fixes.fix_except import find_excepts
from multiprocessing import Value
from pdb import find_function
from re import L
from turtle import color
import pandas as pd
import numpy as np
from io import BytesIO
from flask import send_file
import os


from .cdataRefs import *




###### HELPER FUNCTIONS ######

import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule, Rule, DifferentialStyle
from openpyxl import formatting, styles
from openpyxl.styles import (
    Color,
    PatternFill,
    Font,
    Fill,
    Border,
    Side,
    Alignment,
    GradientFill,
)
from openpyxl.drawing.image import Image
from openpyxl.worksheet.hyperlink import Hyperlink

# styling functions for worksheets

# header styling
grey_fill = PatternFill(end_color="d0cece", start_color="d0cece", fill_type="solid")

def FindEndCol(sheet,analyte_row,startCol):
    EndCol=startCol
    link=sheet
    j=analyte_row
    ReachEnd=False
    while ReachEnd==False:
        t=get_column_letter(EndCol)
#        print(t,j)
        if link[f"{t}{j}"].value==None:
            ReachEnd=True
#        if len(v)>0:
#            EndCol+=1
        else:
            EndCol+=1
#            ReachEnd==True
    return EndCol

def GetColLetter(sheet,analyte,analyte_row,startCol,endCol):
 #   print("GetColLetter")
 #   print(analyte +"analyte_row:"+str(analyte_row)+"startcol:"+str(startCol)+"endcol:"+str(endCol))
          
#    print(analyte_row)
#    print(startCol)
#    print("number of column: " + str(endCol-startCol))
    
    
    t_col_letter=""
    i=startCol
    link=sheet
    j=analyte_row
    foundFlag=False
    
    while foundFlag==False:
        while i < endCol:
            t_col_letter=get_column_letter(i)
#            print(t_col_letter)
            checkAnalyte=link[f"{t_col_letter}{j}"].value
#            print(analyte)
#           if(analyte=='Toluene'):
#                print(checkAnalyte)
            if analyte==checkAnalyte:
                foundFlag=True
                i=endCol+1
 #               print(analyte + ":"+checkAnalyte+":"+t_col_letter)
            else:
                i=i+1
#    print('Return' + t_col_letter)
    return t_col_letter

def set_grey_fill(ws, cell_range):
    grey_fill = PatternFill(end_color="d0cece", start_color="d0cece", fill_type="solid")
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.fill = grey_fill

def set_table1_header_format(ws, cell_range):
    blue_fill = PatternFill(end_color = "DEEDF2", start_color="DEEDF2", fill_type = "solid")
    blue_font = styles.Font(size=11, underline='single', italic=True, color='1738E3') 
    border = Border(
            top=thin_black_border,
            bottom=thin_black_border)
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.fill = blue_fill
            cell.font = blue_font
            cell.border = border
            
def set_header_format(ws, cell_range):
    blue_fill = PatternFill(end_color = "FFFFD1", start_color="FFFFD1", fill_type = "solid")
    blue_font = styles.Font(size=11, bold=True,  italic=True, color='1738E3')
    border = Border(
                top=thin_black_border,
                bottom=thin_black_border)
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.fill = blue_fill
            cell.font = blue_font
            cell.border = border

# blue and bold table 1 values
def set_check_font(ws, cell_range):
    checkmark_font = styles.Font(size=11, bold=True, color="1738E3")
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.font = checkmark_font

# center all values in cells
def center_cell(ws, cell_range):
    alignment = Alignment(horizontal="center", vertical="center")
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.alignment = alignment


# center TO THE RIGHT all values in cells
def center_right_cell(ws, cell_range):
    alignment = Alignment(horizontal="right")
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.alignment = alignment


# add black border to all cells
thin_black_border = Side(border_style="thin", color="000000")


def set_border(ws, cell_range):
    border = Border(
        left=thin_black_border,
        right=thin_black_border,
        top=thin_black_border,
        bottom=thin_black_border,
    )

    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.border = border


# format alternating rows background
def alt_row_color(cell_range, first_crit_loc, sheet):
    green_fill = styles.PatternFill(
        end_color="DAE4BF", start_color="DAE4BF", fill_type="solid"
    )
    sheet.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'${first_crit_loc}="g"'], stopIfTrue=False, fill=green_fill
        ),
    )


def old_alt_row_color(cell_range, sheet):
    green_fill = styles.PatternFill(
        end_color="e2efda", start_color="e2efda", fill_type="solid"
    )
    sheet.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=["MOD(ROW(),2)=0"], stopIfTrue=False, fill=green_fill),
    )


# create a column that has alternating "g" and "w" so user can edit background color in excel
def alt_numbers_col(color_col, start_row, stop_row, sheet):
    for row in sheet.iter_rows(
        min_row=start_row, max_row=stop_row, min_col=color_col, max_col=color_col
    ):
        for cell in row:
            cell.value = (
                f'=IF(ISODD(ROW({cell.column_letter + str(cell.row)})),"g","w")'
            )
 #           print(cell.value)


# def sort_results_col(ordering_col, start_row, stop_row, sheet):
#     for row in sheet.iter_rows(min_row=start_row, max_row=stop_row, min_col=ordering_col, max_col=ordering_col):
#         for cell in row:
#             cell.value = f'=IF(ISODD(ROW({cell.column_letter + str(cell.row)})),"w","g")'


def createList(crit_len):
    return [item for item in range(1, crit_len + 1)]


def set_col_width(sheet):
    dims = {}
    for row in sheet.iter_rows(min_row=None, max_row=None, min_col=5, max_col=None):
        for cell in row:
            if cell.value:
                # wrap is cell is too long
                if len(str(cell.value)) > 12:
                    cell.alignment = Alignment(
                        wrapText=True, horizontal="center", vertical="center"
                    )
                # center and space if not too long
                if len(str(cell.value)) <= 12:
                    dims[cell.column_letter] = max(
                        (dims.get(cell.column_letter, 4), len(str(cell.value)))
                    )
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value + 2


def adjust_col_width(sheet):
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value)))
                )
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value + 2


def wrap_txt(sheet):
    for rows in sheet.iter_rows(min_row=1, max_row=1, min_col=6, max_col=None):
        for cell in rows:
            if len(str(cell.value)) > 12:
                cell.alignment = Alignment(
                    wrapText=True, horizontal="center", vertical="center"
                )


def wrap_headers(sheet, header_row, max_cols, max_cell_length):
    for rows in sheet.iter_rows(
        min_row=header_row, max_row=header_row, min_col=1, max_col=max_cols
    ):
        for cell in rows:
            if len(str(cell.value)) > max_cell_length:
                cell.alignment = Alignment(
                    wrapText=True, horizontal="center", vertical="center"
                )


